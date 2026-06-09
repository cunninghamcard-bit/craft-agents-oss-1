#!/usr/bin/env python3
"""把一张订单(JSON)填进「美金请款发票 PI」模板，输出一份可编辑 .xlsx。

美金请款发票 PI（USD Proforma Invoice）专用渲染器。读 JSON context → 填模板 Sheet1 →
保留模板版式/公式（金额=单价×数量、合计 SUM 自动算）→ 货品行数不固定时插行并顺延公式 →
存成新 .xlsx（不改模板本身）。

用法（openpyxl）：
  uv run --with openpyxl python3 render_pi.py \
      --template templates/美金请款发票模板PI.xlsx \
      --out "PI_<订单号>.xlsx" \
      --data ctx.json
  # --data 也可走 stdin

ctx.json 形状：
  {
    "invoice_no": "8828984",            # Invoice #（H7）
    "invoice_date": "20260609",         # 日期（B7 -> "Date:"+值）
    "bill_to": {"name":"...","address":"...","tel":"..."},
    "ship_to": {"name":"...","address":"...","tel":"..."},   # 缺省=同 bill_to
    "po_number": "8828984", "terms":"DDU",
    "ship_date":"", "ship_via":"", "tracking_no":"",
    "items": [ {"part":"...","qty":900,"desc":"...","price":37}, ... ]
  }
成功时把输出路径打到 stdout。
"""
import argparse
import copy
import json
import os
import sys

TPL_SHEET = "Sheet1"          # 模板里用作母版的 sheet（其余历史样例 sheet 会被丢弃）
ITEM_START = 17               # 货品首行
TPL_ITEM_ROWS = 2             # 模板自带 2 个货品行（17、18）
TPL_TOTAL_ROW = 19            # 模板里的合计行


def _set(ws, coord, value):
    ws[coord] = value


def _copy_row_style(ws, src_row, dst_row, max_col):
    """把 src_row 的单元格样式/行高复制到 dst_row（insert_rows 出来的行是裸的）。"""
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if s.has_style:
            d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill)
            d.border = copy.copy(s.border)
            d.alignment = copy.copy(s.alignment)
            d.number_format = s.number_format
            d.protection = copy.copy(s.protection)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _merge_desc(ws, row):
    """货品行描述列 E:F 合并（模板每个货品行都是 E:F 合并）。"""
    rng = f"E{row}:F{row}"
    try:
        ws.merge_cells(rng)
    except Exception:
        pass


def _insert_rows_keep_merges(ws, idx, amount):
    """openpyxl 的 insert_rows 不会下移合并区，会把上一块的合并残留到新行、
    让被合并的非锚格写入失效。这里手动把 idx 行及其下的合并区整体下移 amount 行。"""
    from openpyxl.utils import range_boundaries
    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr, maxc, maxr))
    ws.insert_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(start_row=minr + amount, start_column=minc,
                       end_row=maxr + amount, end_column=maxc)


def _delete_rows_keep_merges(ws, idx, amount):
    """对称地处理删行：删掉 [idx, idx+amount) 的合并区，其下的合并区上移。"""
    from openpyxl.utils import range_boundaries
    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx + amount:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr - amount, maxc, maxr - amount))
        elif minr >= idx:
            ws.unmerge_cells(str(rng))  # 落在被删区间内，丢弃
    ws.delete_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(start_row=minr, start_column=minc, end_row=maxr, end_column=maxc)


def main() -> None:
    ap = argparse.ArgumentParser(description="Render the USD PI (美金请款发票) from order JSON.")
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--data", help="JSON context 文件；省略则读 stdin")
    args = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl 未安装。请用：uv run --with openpyxl python3 <本脚本> ...")

    if not os.path.exists(args.template):
        sys.exit(f"模板不存在：{args.template}")
    raw = open(args.data, encoding="utf-8").read() if args.data else sys.stdin.read()
    if not raw.strip():
        sys.exit("没有数据（--data 或 stdin 给 JSON）")
    try:
        ctx = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"JSON 解析失败：{e}")

    items = ctx.get("items") or []
    if not items:
        sys.exit("items 为空：一张 PI 至少要有一个货品行")

    wb = load_workbook(args.template)
    if TPL_SHEET not in wb.sheetnames:
        sys.exit(f"模板缺少母版 sheet「{TPL_SHEET}」")
    ws = wb[TPL_SHEET]
    # 只留母版，丢弃其余历史样例 sheet
    for name in list(wb.sheetnames):
        if name != TPL_SHEET:
            del wb[name]

    max_col = ws.max_column

    # ── 调整货品行数（插/删行，让货品区正好 = len(items) 行）──
    n = len(items)
    if n > TPL_ITEM_ROWS:
        extra = n - TPL_ITEM_ROWS
        _insert_rows_keep_merges(ws, TPL_TOTAL_ROW, extra)   # 在合计行前插，合计/银行块随之下移
        for i in range(extra):
            dst = ITEM_START + TPL_ITEM_ROWS + i              # 19, 20, ...
            _copy_row_style(ws, ITEM_START, dst, max_col)
            _merge_desc(ws, dst)
    elif n < TPL_ITEM_ROWS:
        _delete_rows_keep_merges(ws, ITEM_START + n, TPL_ITEM_ROWS - n)  # 货品行比模板少，删多余

    total_row = ITEM_START + n

    # ── 抬头 ──
    _set(ws, "B7", f"Date:{ctx.get('invoice_date','')}")
    _set(ws, "H7", ctx.get("invoice_no", ""))
    bt = ctx.get("bill_to") or {}
    st = ctx.get("ship_to") or bt
    _set(ws, "B9", bt.get("name", ""));  _set(ws, "B10", bt.get("address", ""));  _set(ws, "B11", bt.get("tel", ""))
    _set(ws, "F9", st.get("name", ""));  _set(ws, "F10", st.get("address", ""));  _set(ws, "F11", st.get("tel", ""))
    _set(ws, "B14", ctx.get("po_number", ""))
    _set(ws, "D14", ctx.get("terms", ""))
    _set(ws, "E14", ctx.get("ship_date", ""))
    _set(ws, "G14", ctx.get("ship_via", ""))
    _set(ws, "H14", ctx.get("tracking_no", ""))

    # ── 货品行：B 序号 / C 料号 / D 数量 / E 描述 / G 单价 / H=G*D（公式）──
    for i, it in enumerate(items):
        r = ITEM_START + i
        _set(ws, f"B{r}", i + 1)
        _set(ws, f"C{r}", it.get("part", ""))
        _set(ws, f"D{r}", it.get("qty"))
        _set(ws, f"E{r}", it.get("desc", ""))
        _set(ws, f"G{r}", it.get("price"))
        _set(ws, f"H{r}", f"=G{r}*D{r}")

    # ── 合计行：数量合计 / 金额合计（重写 SUM 范围，因为插行不会自动改公式）──
    _set(ws, f"D{total_row}", f"=SUM(D{ITEM_START}:D{total_row-1})")
    _set(ws, f"H{total_row}", f"=SUM(H{ITEM_START}:H{total_row-1})")

    out_dir = os.path.dirname(os.path.abspath(args.out)) or "."
    os.makedirs(out_dir, exist_ok=True)
    wb.save(args.out)
    print(args.out)


if __name__ == "__main__":
    main()
